import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime
from dateutil.relativedelta import relativedelta
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook

def getMedianVolume(data):
	return(round(data.median(),0))

def getDailyReturns(data):
	return(data.pct_change(fill_method = 'ffill')) #Modified

# change inf value to max and min value of that column
def getMaskDailyChange(data) :
 m1 = getDailyReturns(data).eq(np.inf) # for inf value
 m2 = getDailyReturns(data).eq(-np.inf) #for -inf value
 return(getDailyReturns(data).mask(m1, df[~m1].max(), axis=1).mask(m2, df[~m2].min(), axis=1).bfill(axis = 1))


def getStdev(data):
	return(np.std(getMaskDailyChange(data)*100))

def getStdRatio(data, data1):
	return((getStdev(data)/getStdev(data1)*100)) # to judge volatility 1 month against 1 year

def getAbsReturns(data):
	x = (data.iloc[-1]/data.iloc[0] - 1)*100
	return(round(x, 2))

def getVolatility(data):
	return(round(np.std(data) * np.sqrt(252) * 100, 2))

def getMonthlyPrices(data):
	grps = data.groupby([data.index.year, data.index.month])
	monthlyPrices = pd.DataFrame()
	for k in grps:
		monthlyPrices = pd.concat([monthlyPrices, k[1].tail(1)])
		# monthlyPrices = monthlyPrices.append(k[1].tail(1))
	return monthlyPrices

def getMonthlyReturns(data):
	return(data.pct_change())

def getSharpe(data):
	return(round(np.sqrt(252) * data.mean()/data.std(), 2))

def getSortino(data):
	return(np.sqrt(252) * data.mean()/data[data<0].std())

def getMaxDrawdown(data):
	cummRet = (data+1).cumprod()
	peak = cummRet.expanding(min_periods = 1).max()
	drawdown = (cummRet/peak) - 1
	return drawdown.min()

def getCalmar(data):
	return(data.mean()*252/abs(getMaxDrawdown(data)))

def getAbsMomentumVolAdjusted(absReturn, volatility):
	return(absReturn/volatility)

def getNMonthRoC(data, N):
	ret = round((data.iloc[-1]/data.iloc[-1-N] - 1) * 100, 2)
	return(ret)

def getNWeekRoC(data, N):
	ret = round((data.iloc[-1]/data.iloc[-1-N] - 1) * 100, 2)
	return(ret)

def getFIP(data):
	retPos = np.sum(data.pct_change()[1:] > 0)
	retNeg = np.sum(data.pct_change()[1:] < 0)
	return(retPos - retNeg)

def getSharpeRoC(roc, volatility):
	return(round(roc/volatility, 2))

#Beta should be calculated against relevant Index instead of Nifty50?
def getBeta(dfNifty, data12M):

	dailyReturns = getDailyReturns(pd.concat([dfNifty, data12M], axis = 1))[1:]

	var = np.var(dailyReturns['Nifty']) #Modified

	cov = dailyReturns.cov()

	cols = cov.columns[1:]

	beta = []

	for k in cols:
		beta.append(round(cov.loc[k, 'Nifty']/var, 2))

	return beta

# Streamlit Layout
st.title("ETF Momentum Ranking App")


#To suppress future warnings about using ffill method in pct_change()
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

# Dropdown options with display labels and corresponding values
ranking_options = {
    "Sharpe3M": "sharpe3M",
    "AvgSharpe 12M/9M/6M/3M": "avgSharpe"
}

# Display dropdown for ranking method selection
ranking_method_display = st.selectbox(
    "Select Ranking Method",
    options=list(ranking_options.keys()),  # Display labels
    index=0  # Default to the first option
)

# Get the actual value for the selected display label
ranking_method = ranking_options[ranking_method_display]

# Select Universe with default value as 'NSEETF'
universe = ['NSEETF']
U = st.selectbox('Select Universe:', universe, index=0)  # Default value is 'NSEETF' (index 0)

# Date Picker for Lookback Start Date
selected_date = st.date_input("Select Lookback Date", datetime.today())
dt2 = datetime.strptime(str(selected_date), "%Y-%m-%d").strftime('%Y-%m-%d')

# Displaying Date Range Information
dates = {
    'startDate': datetime.strptime('2000-01-01', '%Y-%m-%d'),
    'endDate': datetime.strptime(dt2, '%Y-%m-%d'),
    'date12M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=12),
    'date9M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=9),
    'date6M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=6),
    'date3M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=3),
    'date1M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=1),
}

st.write("##### Date Range:")
st.write(f"Start Date: **{dates['startDate'].strftime('%d-%m-%Y')}**")
st.write(f"End Date: **{dates['endDate'].strftime('%d-%m-%Y')}**")


# Read index file based on selected universe
if U == 'NSEETF':
    file_path = 'https://raw.githubusercontent.com/prayan2702/ETF-Momo-app/refs/heads/main/NSE_ETF.csv'


df = pd.read_csv(file_path)
df['Yahoo_Symbol'] = df.Symbol + '.NS'
df = df.set_index('Yahoo_Symbol')
symbol = list(df.index)

# Add a button to start the process
start_button = st.button("Start Data Download")

if start_button:
    # Download data when the button is pressed
    CHUNK = 50
    close = []
    high = []
    volume = []

    # Create a progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()  # Placeholder for progress text

    # Track the number of stocks downloaded
    total_symbols = len(symbol)
    chunk_count = (total_symbols // CHUNK) + (1 if total_symbols % CHUNK != 0 else 0)

    # Use a separate list to handle failed downloads
    failed_symbols = []

    for k in range(0, len(symbol), CHUNK):
        _symlist = symbol[k:k + CHUNK]

        # Try downloading data for each chunk of symbols
        try:
            _x = yf.download(_symlist, start=dates['startDate'], progress=False, auto_adjust = True, threads = True, multi_level_index=False)
            close.append(_x['Close'])
            high.append(_x['High'])
            volume.append(_x['Close'] * _x['Volume'])
        except Exception as e:
            failed_symbols.extend(_symlist)  # Add failed symbols to the list
            st.write(f"Failed to download data for: {', '.join(_symlist)}. Error: {e}")

        # Update progress bar after each chunk
        progress = (k + CHUNK) / total_symbols
        progress = min(max(progress, 0.0), 1.0)  #newly added for progress bar error
        progress_bar.progress(progress)

        # Update status text with progress percentage
        progress_percentage = int(progress * 100)
        status_text.text(f"Downloading... {progress_percentage}%")

        time.sleep(0.5)

        # After the download is complete, update the progress bar and text
    progress_bar.progress(1.0)
    status_text.text("Download complete!")

#*******************************************
    # Applied filter descriptions
    filters = [
        "Volume greater than 1 crore (volm_cr > 1)",
        "Closing price above 200-day moving average (Close > dma200d)",
        "12-month Rate of Change (ROC) greater than 6.5% (roc12M > 6.5)",
        "Away from All-Time High within 25% (AWAY_ATH > -25)"
    ]

    # Sidebar menu for filters
    with st.sidebar:
        st.header("Filters Menu")
        with st.expander("Applied Filters", expanded=False):
            st.write("The following conditions are applied:")
            for i, filter_desc in enumerate(filters, start=1):
                st.write(f"{i}. {filter_desc}")

#**************************************************


    # Handle failed symbols (if any)
    if failed_symbols:
        st.write(f"Failed to download data for the following symbols: {', '.join(failed_symbols)}")


    # Convert close, high, and volume lists to DataFrames
    close = pd.concat(close, axis=1) if close else pd.DataFrame()
    high = pd.concat(high, axis=1) if high else pd.DataFrame()
    volume = pd.concat(volume, axis=1) if volume else pd.DataFrame()

    # Ensure the index is datetime
    close.index = pd.to_datetime(close.index)
    high.index = pd.to_datetime(high.index)
    volume.index = pd.to_datetime(volume.index)

    data20Y = close.loc[:dates['endDate']].copy()
    volume20Y = volume.loc[:dates['endDate']].copy()
    high20Y = high.loc[:dates['endDate']].copy()
    volume12M = volume20Y.loc[dates['date12M']:].copy()


    # At least 12 months of trading is required
    data12M = data20Y.loc[dates['date12M']:].copy()
    data9M = data20Y.loc[dates['date9M']:].copy()
    data6M = data20Y.loc[dates['date6M']:].copy()
    data3M = data20Y.loc[dates['date3M']:].copy()
    data1M = data20Y.loc[dates['date1M']:].copy()

    #******************

    # Calculate metrics for dfStats (e.g., 1 month momentum, 3-month volatility, etc.)
    dfStats = pd.DataFrame(index=symbol)
    dfStats['Close'] = round(data12M.iloc[-1], 2)
    data12M_Temp = data12M.fillna(0)
    dfStats['dma200d'] = round(data12M_Temp.rolling(window=200).mean().iloc[-1], 2)  # 200-day DMA
    # Rate of change

    dfStats['roc12M'] = getAbsReturns(data12M)
    dfStats['roc9M'] = getAbsReturns(data9M)
    dfStats['roc6M'] = getAbsReturns(data6M)
    dfStats['roc3M'] = getAbsReturns(data3M)


    # Volatility

    dfStats['vol12M'] = getVolatility(getDailyReturns(data12M))
    dfStats['vol9M'] = getVolatility(getDailyReturns(data9M))
    dfStats['vol6M'] = getVolatility(getDailyReturns(data6M))
    dfStats['vol3M'] = getVolatility(getDailyReturns(data3M))

    dfStats['sharpe12M'] = getSharpeRoC(dfStats['roc12M'], dfStats['vol12M'])
    dfStats['sharpe9M'] = getSharpeRoC(dfStats['roc9M'], dfStats['vol9M'])
    dfStats['sharpe6M'] = getSharpeRoC(dfStats['roc6M'], dfStats['vol6M'])
    dfStats['sharpe3M'] = getSharpeRoC(dfStats['roc3M'], dfStats['vol3M'])

    dfStats['avgSharpe'] = (dfStats[["sharpe12M", "sharpe9M", "sharpe6M", "sharpe3M"]].mean(axis=1)).round(2)  # 1st Factor #1st Factor

    dfStats['volm_cr'] = (getMedianVolume(volume12M) / 1e7).round(2)

    #***************************

    # Calculate ATH and Away from ATH% (Additional added condition)
    dfStats['ATH'] = round(high20Y.max(), 2)
    dfStats['AWAY_ATH'] = round((dfStats['Close'] / dfStats['ATH'] - 1) * 100, 2)  # Calculate %away from ALL TIME HIGH


    # Add 'Ticker' as a reset index column and rename it
    dfStats = dfStats.reset_index().rename(columns={'index': 'Ticker'})

    #Convert to String and Remove the .NS Suffix
    dfStats['Ticker'] = dfStats['Ticker'].astype(str)
    dfStats['Ticker'] = dfStats['Ticker'].str.replace('.NS', '', case=False, regex=False)

    # Handle Nan and inf values to zero(0) for ranking
    dfStats['avgSharpe'] = dfStats['avgSharpe'].replace([np.inf, -np.inf], np.nan).fillna(0)
    dfStats['sharpe3M'] = dfStats['sharpe3M'].replace([np.inf, -np.inf], np.nan).fillna(0)


    # Add Rank column based on 'avgSharpe' and sort by Rank
    dfStats['Rank'] = dfStats[ranking_method].rank(ascending=False,method='first').astype(int)
    dfStats = dfStats.sort_values('Rank').set_index('Rank')  # Set 'Rank' as index

    # Show both filtered and unfiltered data in Streamlit
    st.info("Unfiltered Data:")
    st.write(dfStats)

    # Apply conditions
    cond1 = dfStats['volm_cr'] > 1  #volume filter greater than 1 crore
    cond2 = dfStats['Close'] > dfStats['dma200d']  # above 200-day DMA
    cond3 = dfStats['roc12M'] > 6.5  # 12 month Rate of Change > 6.5%
    cond4 = dfStats['AWAY_ATH'] > -25 #Away from All Time High within 25%


    # Create final momentum filter column
    dfStats['final_momentum'] = cond1 & cond2 & cond3 & cond4


#*************************************

    # Filter stocks meeting all conditions
    filtered = dfStats[dfStats['final_momentum']].sort_values(ranking_method, ascending=False)

    st.info("Filtered Data:")
    st.write(filtered)

#***********************************************************
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


    def format_excel(file_name):
        # Open the written file using openpyxl
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

        # Add Borders to All Cells
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                             bottom=Side(style="thin"))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze the top row
        ws.freeze_panes = 'A2'

        # Format headers
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Automatically adjust column widths based on content
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[cell.column_letter].width = adjusted_width

        # Define cell color for cells that do not meet filter conditions
        no_condition_fill = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
        bold_font = Font(bold=True)

        # Get the headers and find column indexes by name
        headers = [cell.value for cell in ws[1]]
        col_indices = {
            'volm_cr': headers.index('volm_cr') + 1,
            'Close': headers.index('Close') + 1,
            'dma200d': headers.index('dma200d') + 1,
            'AWAY_ATH': headers.index('AWAY_ATH') + 1,
            'roc12M': headers.index('roc12M') + 1,
            'Ticker': headers.index('Ticker') + 1,
            'Rank': headers.index('Rank') + 1
        }

        # Apply conditional formatting
        for row in range(2, ws.max_row + 1):
            condition_failed = False
            if (volume := ws.cell(row=row, column=col_indices['volm_cr']).value) is not None and volume < 1:
                ws.cell(row=row, column=col_indices['volm_cr']).fill = no_condition_fill
                condition_failed = True
            if (close := ws.cell(row=row, column=col_indices['Close']).value) is not None and close <= ws.cell(row=row,
                                                                                                               column=
                                                                                                               col_indices[
                                                                                                                   'dma200d']).value:
                ws.cell(row=row, column=col_indices['Close']).fill = no_condition_fill
                condition_failed = True
            if (away_ath := ws.cell(row=row, column=col_indices['AWAY_ATH']).value) is not None and away_ath <= -25:
                ws.cell(row=row, column=col_indices['AWAY_ATH']).fill = no_condition_fill
                condition_failed = True
            if (roc12M := ws.cell(row=row, column=col_indices['roc12M']).value) is not None and roc12M <= 6.5:
                ws.cell(row=row, column=col_indices['roc12M']).fill = no_condition_fill
                condition_failed = True
            if condition_failed:
                ws.cell(row=row, column=col_indices['Ticker']).fill = no_condition_fill

                # Round off "ATH" column values
                ath_idx = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "ATH":  # Search for "ATH" header
                        ath_idx = col
                        break
                if ath_idx:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=ath_idx)
                        if isinstance(cell.value, (int, float)):
                            cell.value = round(cell.value)

        # Save the modified Excel file
        wb.save(file_name)
        print(f"\nExcel file '{file_name}' updated with formatting\n")


#*********************************************************


    def format_filtered_excel(file_name):
        # Open the written file using openpyxl
        wb = openpyxl.load_workbook(file_name)
        ws = wb["Filtered ETF"]  # Specify the "Filtered Stocks" sheet

        # Add Borders to All Cells
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                             bottom=Side(style="thin"))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze the top row
        ws.freeze_panes = 'A2'

        # Format headers
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Automatically adjust column widths based on content
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Round off "ATH" column values
        ath_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "ATH":  # Search for "ATH" header
                ath_idx = col
                break
        if ath_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=ath_idx)
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value)

        # Append '%' to "AWAY_ATH" column values
        away_ath_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "AWAY_ATH":
                away_ath_idx = col
                break

        if away_ath_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=away_ath_idx)
                if isinstance(cell.value, (int, float)):
                    cell.value = f"{cell.value}%"

	 # Add summary
        total_filtered_stocks = ws.max_row - 1
        ws.append([])  # Empty row
        ws.append(["Summary"])  # Summary heading
        summary_start_row = ws.max_row
        ws.append([f"Total Filtered ETF:  {total_filtered_stocks}"])


        # Apply bold font to the summary
        for row in ws.iter_rows(min_row=summary_start_row, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        wb.save(file_name)
        print("\nFiltered Excel file formatted and updated with summary.\n")



#********************************************************
    # Format the filename with the lookback date, universe, and other parameters
    excel_file = f"{selected_date.strftime('%Y-%m-%d')}_{U}_{ranking_method}_lookback.xlsx"

    # Save filtered data to Excel
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        dfStats.to_excel(writer, sheet_name="Unfiltered ETF", index=True)  # Unfiltered data
        filtered.to_excel(writer, sheet_name="Filtered ETF", index=True)  # Filtered data

    # Format the Unfiltered Excel file
    format_excel(excel_file)
    # Format the filtered sheet
    format_filtered_excel(excel_file)

    # Download button for the Excel file
    st.download_button(
        label="Download Stock Data as Excel",
        data=open(excel_file, "rb").read(),
        file_name=excel_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
#***************************************************************

